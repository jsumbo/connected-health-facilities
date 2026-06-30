"""
Load TRIBE master facility readiness scores from the workbook in repo.
Sheets used: 7 Facility Scorecards, 8 County Summary, 9 Cluster Summary, 10 Blocker Register.
"""

from __future__ import annotations

import logging
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, TypedDict

from openpyxl import load_workbook

from drf import (
    build_drf_domain_scores,
    enrich_blockers,
    normalize_facility_name,
    parse_blocker_codes,
    tier_display_label,
)
from facility_master import all_programme_facilities, cluster_sort_key, registry_by_slug

logger = logging.getLogger(__name__)

SHEET_SCORECARDS = "7. Facility Scorecards"
SHEET_COUNTY = "8. County Summary"
SHEET_CLUSTER = "9. Cluster Summary"
SHEET_BLOCKERS = "10. Blocker Register"

DOMAIN_COLS = {
    "D_POW": 5,
    "D_CON": 6,
    "D_ICT": 7,
    "D_DIG": 8,
    "D_SEN": 9,
    "D_DAT": 10,
}


class ScorecardColumnMap(TypedDict):
    rank: int
    facility: int
    county: int
    facility_type: int
    composite: int
    tier: int
    wave: int
    blockers: int
    blockers_exp: int | None
    dla_pct: int | None
    dla_n: int | None
    sentiment_n: int | None


class MasterScorecard(TypedDict):
    slug: str
    facility_name: str
    county: str
    cluster: str
    facility_type: str
    rank: int | None
    composite: float
    tier: str
    tier_label: str
    wave: str | None
    deployment_wave: str | None
    blocker_codes: List[str]
    blockers: List[Dict[str, str]]
    blocker_remediation: str | None
    deployment_blocked: bool
    domain_scores: Dict[str, Dict[str, Any]]
    dla_pct: float | None
    dla_n: int | None
    sentiment_n: int | None
    scoring_source: str


class MasterReadinessBundle(TypedDict):
    scorecards: Dict[str, MasterScorecard]
    blocker_register: List[Dict[str, Any]]
    county_summaries: List[Dict[str, Any]]
    cluster_summaries: List[Dict[str, Any]]
    source_path: str
    facility_count: int


def _default_workbook_path() -> Path:
    return Path(__file__).resolve().parent / "data" / "master" / "Master Facility Readiness Scores.xlsx"


def _build_name_to_slug() -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for reg in all_programme_facilities():
        mapping[normalize_facility_name(reg["name"]).lower()] = reg["slug"]
    return mapping


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _row_has_value(cells: List[Any], index: int) -> bool:
    return index < len(cells) and cells[index] not in (None, "")


def _cell_str(cells: List[Any], index: int) -> str:
    if not _row_has_value(cells, index):
        return ""
    return str(cells[index]).strip()


def _row_cells(row: tuple) -> List[Any]:
    return list(row)


def _is_blank_summary_row(cells: List[Any]) -> bool:
    if not cells:
        return True
    first = cells[0]
    return first is None or str(first).strip() == ""


def _resolve_slug(facility_name: str, name_to_slug: Dict[str, str]) -> str | None:
    key = normalize_facility_name(facility_name).lower()
    return name_to_slug.get(key)


def _normalize_header_cell(value: Any) -> str:
    return " ".join(str(value or "").split()).strip().lower()


def _resolve_scorecard_columns(header_row: tuple) -> ScorecardColumnMap:
    """Map scorecard sheet headers to column indices (supports v1 and v2 layouts)."""
    by_header: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _normalize_header_cell(cell)
        if key:
            by_header[key] = idx

    def col(*candidates: str) -> int | None:
        for name in candidates:
            if name in by_header:
                return by_header[name]
        return None

    def require(*candidates: str) -> int:
        found = col(*candidates)
        if found is None:
            raise ValueError(f"Missing scorecard column: {candidates[0]}")
        return found

    return ScorecardColumnMap(
        rank=require("rank"),
        facility=require("facility"),
        county=require("county"),
        facility_type=require("type"),
        composite=require("composite"),
        tier=require("tier"),
        wave=require("wave"),
        blockers=require("blockers"),
        blockers_exp=col("blockers exp"),
        dla_pct=col("dla %"),
        dla_n=col("dla n"),
        sentiment_n=col("sent n"),
    )


def _cell_at(cells: List[Any], index: int | None) -> Any:
    if index is None or index >= len(cells):
        return None
    return cells[index]


def _domain_score_value(scorecard: MasterScorecard, domain_key: str) -> float | None:
    domain = (scorecard.get("domain_scores") or {}).get(domain_key)
    if not isinstance(domain, dict):
        return None
    score = domain.get("score")
    return float(score) if score is not None else None


def _aggregate_cluster_summaries(
    scorecards: Dict[str, MasterScorecard],
) -> List[Dict[str, Any]]:
    """Roll up cluster KPIs from facility scorecards (registry cluster assignment)."""
    by_cluster: Dict[str, List[MasterScorecard]] = defaultdict(list)
    for scorecard in scorecards.values():
        by_cluster[scorecard["cluster"]].append(scorecard)

    summaries: List[Dict[str, Any]] = []
    for cluster in sorted(by_cluster.keys(), key=cluster_sort_key):
        items = by_cluster[cluster]
        composites = [float(item["composite"]) for item in items if item.get("composite") is not None]
        tier_1_count = sum(1 for item in items if str(item.get("tier", "")).startswith("Tier 1"))
        tier_2_count = sum(1 for item in items if str(item.get("tier", "")).startswith("Tier 2"))
        tier_3_count = sum(1 for item in items if str(item.get("tier", "")).startswith("Tier 3"))

        domain_avg_fields = {
            "D_POW": "avg_pow",
            "D_CON": "avg_con",
            "D_ICT": "avg_ict",
            "D_DIG": "avg_dig",
            "D_SEN": "avg_sen",
            "D_DAT": "avg_dat",
        }
        domain_avgs: Dict[str, float | None] = {}
        for domain_key, field in domain_avg_fields.items():
            vals = [
                value
                for item in items
                if (value := _domain_score_value(item, domain_key)) is not None
            ]
            domain_avgs[field] = round(sum(vals) / len(vals), 2) if vals else None

        summaries.append({
            "cluster": cluster,
            "facility_count": len(items),
            "avg_composite": round(sum(composites) / len(composites), 2) if composites else None,
            "tier_1_count": tier_1_count,
            "tier_2_count": tier_2_count,
            "tier_3_count": tier_3_count,
            **domain_avgs,
        })

    return summaries


def load_master_readiness(path: Path | None = None) -> MasterReadinessBundle:
    workbook_path = path or _default_workbook_path()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Master readiness workbook not found: {workbook_path}")

    name_to_slug = _build_name_to_slug()
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    remediation_by_name: Dict[str, str] = {}
    blocker_register: List[Dict[str, Any]] = []

    if SHEET_BLOCKERS in wb.sheetnames:
        for row in wb[SHEET_BLOCKERS].iter_rows(min_row=2, values_only=True):
            cells = _row_cells(row)
            if len(cells) < 6 or not cells[0]:
                continue
            facility_name = str(cells[0]).strip()
            slug = _resolve_slug(facility_name, name_to_slug)
            codes = parse_blocker_codes(str(cells[4] or ""))
            remediation = str(cells[5]).strip() if cells[5] else None
            remediation_by_name[normalize_facility_name(facility_name).lower()] = remediation or ""
            blocker_register.append({
                "facility_name": facility_name,
                "facility_slug": slug,
                "county": str(cells[1]).strip() if cells[1] else None,
                "composite": _safe_float(cells[2]),
                "tier": str(cells[3]).strip() if cells[3] else None,
                "blocker_codes": codes,
                "remediation_pathway": remediation,
            })

    scorecards: Dict[str, MasterScorecard] = {}
    if SHEET_SCORECARDS in wb.sheetnames:
        scorecard_ws = wb[SHEET_SCORECARDS]
        header_row = next(scorecard_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ValueError(f"Missing header row in {SHEET_SCORECARDS}")
        scorecard_cols = _resolve_scorecard_columns(header_row)

        for row in scorecard_ws.iter_rows(min_row=2, values_only=True):
            cells = _row_cells(row)
            if len(cells) <= scorecard_cols["composite"] or not _row_has_value(cells, scorecard_cols["facility"]):
                continue

            facility_name = _cell_str(cells, scorecard_cols["facility"])
            slug = _resolve_slug(facility_name, name_to_slug)
            if not slug:
                logger.warning("Master scorecard facility not in registry: %s", facility_name)
                continue

            domain_raw = {
                key: _safe_int(cells[col]) if col < len(cells) else None
                for key, col in DOMAIN_COLS.items()
            }
            wave_raw = _cell_str(cells, scorecard_cols["wave"])
            wave = None if wave_raw in ("", "—", "-") else wave_raw
            tier = _cell_str(cells, scorecard_cols["tier"]) or "Not Assessed"
            blocker_codes = parse_blocker_codes(_cell_str(cells, scorecard_cols["blockers"]))
            remediation = remediation_by_name.get(normalize_facility_name(facility_name).lower())
            if not remediation and blocker_codes:
                remediation = "; ".join(
                    enrich_blockers(blocker_codes)[i]["remediation"]
                    for i in range(len(blocker_codes))
                )

            dla_pct_col = scorecard_cols["dla_pct"]
            dla_n_col = scorecard_cols["dla_n"]
            sentiment_n_col = scorecard_cols["sentiment_n"]

            scorecards[slug] = {
                "slug": slug,
                "facility_name": registry_by_slug()[slug]["name"],
                "county": _cell_str(cells, scorecard_cols["county"]),
                "cluster": registry_by_slug()[slug]["cluster"],
                "facility_type": _cell_str(cells, scorecard_cols["facility_type"]),
                "rank": _safe_int(_cell_at(cells, scorecard_cols["rank"])),
                "composite": _safe_float(_cell_at(cells, scorecard_cols["composite"])) or 0.0,
                "tier": tier,
                "tier_label": tier_display_label(tier, wave),
                "wave": wave,
                "deployment_wave": wave,
                "blocker_codes": blocker_codes,
                "blockers": enrich_blockers(blocker_codes, remediation),
                "blocker_remediation": remediation,
                "deployment_blocked": tier == "Tier 3" or bool(blocker_codes),
                "domain_scores": build_drf_domain_scores(domain_raw),
                "dla_pct": _safe_float(_cell_at(cells, dla_pct_col)),
                "dla_n": _safe_int(_cell_at(cells, dla_n_col)),
                "sentiment_n": _safe_int(_cell_at(cells, sentiment_n_col)),
                "scoring_source": "tribe_master_workbook",
            }

    county_summaries: List[Dict[str, Any]] = []
    if SHEET_COUNTY in wb.sheetnames:
        for row in wb[SHEET_COUNTY].iter_rows(min_row=2, values_only=True):
            cells = _row_cells(row)
            if _is_blank_summary_row(cells):
                continue
            county_name = _cell_str(cells, 0)
            if county_name.upper() == "NATIONAL":
                continue
            county_summaries.append({
                "county": county_name,
                "facility_count": _safe_int(cells[1]),
                "avg_composite": _safe_float(cells[2]),
                "tier_1_count": _safe_int(cells[3]),
                "tier_2_count": _safe_int(cells[4]),
                "tier_3_count": _safe_int(cells[5]),
                "avg_pow": _safe_float(cells[6]),
                "avg_con": _safe_float(cells[7]),
                "avg_ict": _safe_float(cells[8]),
                "avg_dig": _safe_float(cells[9]),
                "avg_sen": _safe_float(cells[10]),
                "avg_dat": _safe_float(cells[11]),
            })

    cluster_summaries = _aggregate_cluster_summaries(scorecards)

    wb.close()

    return {
        "scorecards": scorecards,
        "blocker_register": blocker_register,
        "county_summaries": county_summaries,
        "cluster_summaries": cluster_summaries,
        "source_path": str(workbook_path),
        "facility_count": len(scorecards),
    }
